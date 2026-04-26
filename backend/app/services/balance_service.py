"""Balance P&G data service — extracted from pages/balance.py"""
import io
import os
import re
from typing import List, Tuple, Optional

import duckdb
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH = os.environ.get("DUCKDB_PATH", "/home/juanstn/git/streamlit_congroup/mazda.duckdb")

AZUL = "#1B3A6B"
AZUL_ALT = "#254E9A"
ROJO = "#C0392B"
VERDE = "#0D7C66"
GRIS_TOTAL = "#2D3748"

MESES_LABELS = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


def _con():
    return duckdb.connect(DB_PATH, read_only=True)


def get_filter_options():
    con = _con()
    empresas = con.execute(
        'SELECT DISTINCT empresa FROM mazda.main.dim_empresa ORDER BY 1'
    ).fetchdf()["empresa"].tolist()
    años = con.execute(
        """
        SELECT DISTINCT dt.año
        FROM mazda.main.dim_tiempo dt
        WHERE EXISTS (
            SELECT 1 FROM mazda.main.balance b WHERE CAST(b."Año" AS INT) = dt.año
        )
        ORDER BY 1 DESC
        """
    ).fetchdf()["año"].tolist()
    con.close()
    return empresas, [int(a) for a in años if a is not None]


def _parse_denom_cats(porcentaje: str) -> list:
    if not porcentaje or pd.isna(porcentaje):
        return []
    parts = porcentaje.split("/", 1)
    if len(parts) < 2:
        return []
    return re.findall(r"\b([A-Z][A-Z0-9]*\d+)\b", parts[1])


def _load_and_build(empresa, año):
    con = _con()
    conditions = []
    if empresa and empresa != "Todas":
        conditions.append(f'"Empresa Nombre" = \'{empresa}\'')
    if año and año != "Todos":
        conditions.append(f'"Año" = {int(año)}')
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    df_leaf = con.execute(f"""
        SELECT Clave, Mes, SUM("Saldo Mensual") AS Resultado
        FROM mazda.main.balance
        {where}
        GROUP BY Clave, Mes
    """).fetchdf()

    plantilla = con.execute(
        "SELECT * FROM mazda.main.plantilla_resultado ORDER BY N"
    ).fetchdf()
    con.close()

    if df_leaf.empty:
        return pd.DataFrame(), []

    cat_values: dict[str, dict] = {}
    for _, row in df_leaf.iterrows():
        mes = int(row["Mes"]) if not pd.isna(row["Mes"]) else None
        if mes is None:
            continue
        cat_values.setdefault(row["Clave"], {})[mes] = row["Resultado"]

    meses = sorted({m for cv in cat_values.values() for m in cv})

    for _, t_row in plantilla[
        plantilla["Categoria"].str.startswith("T", na=False)
    ].iterrows():
        t_cat = t_row["Categoria"]
        children = plantilla[plantilla["GrupoPadre"] == t_cat]["Categoria"].tolist()
        cat_values.setdefault(t_cat, {})
        for m in meses:
            cat_values[t_cat][m] = sum(
                cat_values.get(c, {}).get(m, 0) or 0 for c in children
            )

    for cat in cat_values:
        cat_values[cat]["Total"] = sum(cat_values[cat].get(m, 0) or 0 for m in meses)

    cols = meses + ["Total"]
    rows = []
    for _, pr in plantilla.iterrows():
        cat = pr["Categoria"]
        denom_cats = _parse_denom_cats(pr["Porcentaje"])
        vals = cat_values.get(cat, {})
        r = {
            "Categoria": cat,
            "Concepto": pr["Concepto"],
            "is_total": str(cat).startswith("T"),
        }
        for col in cols:
            val = vals.get(col, 0) or 0
            denom = sum(cat_values.get(c, {}).get(col, 0) or 0 for c in denom_cats)
            r[f"res_{col}"] = val
            r[f"pct_{col}"] = (val / denom * 100) if denom != 0 else None
        rows.append(r)

    return pd.DataFrame(rows), meses


def _load_and_build_ppto(empresa, año):
    con = _con()
    empresa_code = None
    if empresa and empresa != "Todas":
        row = con.execute(
            "SELECT code FROM mazda.main.dim_empresa WHERE empresa = ?", [empresa]
        ).fetchone()
        if row:
            empresa_code = int(row[0])

    conditions = []
    if empresa_code is not None:
        conditions.append(f"Empresa = {empresa_code}")
    if año and año != "Todos":
        conditions.append(f"Año = '{año}'")
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    df_leaf = con.execute(f"""
        SELECT Categoria, CAST(Mes AS INT) AS Mes, SUM(TotalBase) AS TotalBase
        FROM mazda.main.ppto_mensual
        {where}
        GROUP BY Categoria, Mes
    """).fetchdf()

    plantilla = con.execute(
        "SELECT * FROM mazda.main.plantilla_resultado ORDER BY N"
    ).fetchdf()
    con.close()

    if df_leaf.empty:
        return pd.DataFrame(), []

    cat_values: dict[str, dict] = {}
    for _, row in df_leaf.iterrows():
        mes = int(row["Mes"])
        cat_values.setdefault(row["Categoria"], {})[mes] = row["TotalBase"]

    meses = sorted({m for cv in cat_values.values() for m in cv})

    for _, t_row in plantilla[
        plantilla["Categoria"].str.startswith("T", na=False)
    ].iterrows():
        t_cat = t_row["Categoria"]
        children = plantilla[plantilla["GrupoPadre"] == t_cat]["Categoria"].tolist()
        cat_values.setdefault(t_cat, {})
        for m in meses:
            cat_values[t_cat][m] = sum(
                cat_values.get(c, {}).get(m, 0) or 0 for c in children
            )

    for cat in cat_values:
        cat_values[cat]["Total"] = sum(cat_values[cat].get(m, 0) or 0 for m in meses)

    cols = meses + ["Total"]
    rows = []
    for _, pr in plantilla.iterrows():
        cat = pr["Categoria"]
        denom_cats = _parse_denom_cats(pr["Porcentaje"])
        vals = cat_values.get(cat, {})
        r = {
            "Categoria": cat,
            "Concepto": pr["Concepto"],
            "is_total": str(cat).startswith("T"),
        }
        for col in cols:
            val = vals.get(col, 0) or 0
            denom = sum(cat_values.get(c, {}).get(col, 0) or 0 for c in denom_cats)
            r[f"ppto_{col}"] = val
            r[f"pct_{col}"] = (val / denom * 100) if denom != 0 else None
        rows.append(r)

    return pd.DataFrame(rows), meses


def _load_and_build_vs_ppto(empresa, año):
    df_real, meses_real = _load_and_build(empresa, año)
    df_ppto, meses_ppto = _load_and_build_ppto(empresa, año)

    if df_real.empty and df_ppto.empty:
        return pd.DataFrame(), []

    meses = sorted(set(meses_real) | set(meses_ppto))
    cols = meses + ["Total"]

    ppto_lookup = {row["Categoria"]: row for _, row in df_ppto.iterrows()} if not df_ppto.empty else {}

    rows = []
    base = df_real if not df_real.empty else df_ppto
    for _, row in base.iterrows():
        cat = row["Categoria"]
        ppto_row = ppto_lookup.get(cat, {})
        r = {
            "Categoria": cat,
            "Concepto": row["Concepto"],
            "is_total": row["is_total"],
        }
        for col in cols:
            real_val = (row.get(f"res_{col}", 0) or 0) if not df_real.empty else 0
            ppto_val = ppto_row.get(f"ppto_{col}", 0) or 0
            diff = real_val - ppto_val
            var_pct = (diff / abs(ppto_val) * 100) if ppto_val != 0 else None
            r[f"real_{col}"] = real_val
            r[f"ppto_{col}"] = ppto_val
            r[f"diff_{col}"] = diff
            r[f"var_{col}"] = var_pct
        rows.append(r)

    return pd.DataFrame(rows), meses


def _fmt(v, pct=False) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if pct:
        return f"{v:.2f} %"
    rounded = int(round(v))
    if rounded < 0:
        return f"({abs(rounded):,})".replace(",", ".")
    return f"{rounded:,}".replace(",", ".")


def _build_excel(df_rows, meses):
    cols = meses + ["Total"]
    export_cols = ["Categoría", "Concepto"]
    for col in cols:
        label = MESES_LABELS.get(col, str(col)).capitalize() if isinstance(col, int) else col
        export_cols.append(f"{label} Resultado")
        export_cols.append(f"{label} %")

    data = []
    for _, row in df_rows.iterrows():
        r = [row["Categoria"], row["Concepto"]]
        for col in cols:
            r.append(row.get(f"res_{col}", 0) or 0)
            r.append(row.get(f"pct_{col}"))
        data.append(r)

    df_export = pd.DataFrame(data, columns=export_cols)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Balance P&G")
        ws = writer.sheets["Balance P&G"]

        header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        total_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        total_font = Font(color="FFFFFF", bold=True, size=11)
        odd_fill = PatternFill(start_color="EEF3FB", end_color="EEF3FB", fill_type="solid")
        odd_stripe_fill = PatternFill(start_color="E4EDF8", end_color="E4EDF8", fill_type="solid")
        tot_col_fill = PatternFill(start_color="E0EAF8", end_color="E0EAF8", fill_type="solid")
        tot_col_stripe_fill = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")
        stripe_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
        red_font = Font(color="C0392B")
        green_font = Font(color="0D7C66")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 38
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col_letter not in ("A", "B"):
                ws.column_dimensions[col_letter].width = 18

        for row_idx, (_, row) in enumerate(df_rows.iterrows()):
            excel_row = row_idx + 2
            is_total = row["is_total"]
            stripe = row_idx % 2 == 1

            for col_idx, cell in enumerate(ws[excel_row], start=1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="right" if col_idx > 2 else "left", vertical="center"
                )
                if is_total:
                    cell.fill = total_fill
                    cell.font = total_font
                    continue
                if col_idx > 2:
                    if col_idx > len(export_cols) - 2:
                        cell.fill = tot_col_stripe_fill if stripe else tot_col_fill
                    else:
                        mes_col_idx = (col_idx - 3) // 2
                        if mes_col_idx % 2 == 1:
                            cell.fill = odd_stripe_fill if stripe else odd_fill
                        else:
                            cell.fill = stripe_fill if stripe else PatternFill(fill_type=None)
                else:
                    cell.fill = stripe_fill if stripe else PatternFill(fill_type=None)
                if col_idx > 2:
                    val = cell.value
                    if val is not None and isinstance(val, (int, float)) and not pd.isna(val):
                        if col_idx % 2 == 0:
                            cell.font = green_font if val > 0 else red_font if val < 0 else Font()
                        else:
                            cell.font = red_font if val < 0 else Font()
            for col_idx in range(3, len(export_cols) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                if col_idx % 2 == 0:
                    cell.number_format = '0.00"%"'
                else:
                    cell.number_format = '#,##0;(#,##0)'
        ws.freeze_panes = "C2"

    buffer.seek(0)
    return buffer.getvalue()


def get_balance_data(empresa: str, año: str):
    df_rows, meses = _load_and_build(empresa, año)
    if df_rows.empty:
        return {"rows": [], "meses": []}
    # Convert NaN to None for JSON serialization
    records = df_rows.replace({pd.NaT: None, float('nan'): None}).to_dict("records")
    return {"rows": records, "meses": meses}


def get_ppto_data(empresa: str, año: str):
    df_rows, meses = _load_and_build_ppto(empresa, año)
    if df_rows.empty:
        return {"rows": [], "meses": []}
    records = df_rows.replace({pd.NaT: None, float('nan'): None}).to_dict("records")
    return {"rows": records, "meses": meses}


def get_vs_ppto_data(empresa: str, año: str):
    df_rows, meses = _load_and_build_vs_ppto(empresa, año)
    if df_rows.empty:
        return {"rows": [], "meses": []}
    records = df_rows.replace({pd.NaT: None, float('nan'): None}).to_dict("records")
    return {"rows": records, "meses": meses}


def get_cuentas_data(empresa: str, año: str, categoria: str):
    con = _con()
    conditions = []
    if empresa and empresa != "Todas":
        conditions.append(f'"Empresa Nombre" = \'{empresa}\'')
    if año and año != "Todos":
        conditions.append(f'"Año" = {int(año)}')
    conditions.append(f"Clave = '{categoria}'")
    where = "WHERE " + " AND ".join(conditions)

    df = con.execute(f'''
        SELECT
            Cuenta,
            "Nombre Cuenta" as nombre_cuenta,
            Nombre as centro_nombre,
            Centro,
            "Sección" as seccion,
            Mes,
            SUM("Saldo Mensual") as Saldo
        FROM mazda.main.balance
        {where}
        GROUP BY Cuenta, "Nombre Cuenta", Nombre, Centro, "Sección", Mes
        ORDER BY ABS(SUM("Saldo Mensual")) DESC
    ''').fetchdf()
    con.close()

    if df.empty:
        return {"cuentas": [], "meses": []}

    meses = sorted([int(m) for m in df["Mes"].dropna().unique().tolist()])

    grouped = df.groupby(["Cuenta", "nombre_cuenta", "centro_nombre", "Centro", "seccion"])
    cuentas = []
    for (cuenta, nombre_cuenta, centro_nombre, centro, seccion), gdf in grouped:
        saldos = {int(row["Mes"]): row["Saldo"] for _, row in gdf.iterrows()}
        total = sum(saldos.values())
        row_data = {
            "cuenta": str(cuenta),
            "nombre_cuenta": nombre_cuenta,
            "centro_nombre": centro_nombre,
            "centro": int(centro) if not pd.isna(centro) else None,
            "seccion": int(seccion) if not pd.isna(seccion) else None,
            "total": total,
        }
        for m in meses:
            row_data[f"saldo_{m}"] = saldos.get(m, 0) or 0
        cuentas.append(row_data)

    cuentas.sort(key=lambda x: abs(x["total"]), reverse=True)
    return {"cuentas": cuentas, "meses": meses}


def get_vs_anterior_data(empresa: str, año: str, mes: int):
    if not año or año == "Todos":
        return {"rows": [], "mes": mes, "error": "Seleccione un año específico"}
    df_actual, meses = _load_and_build(empresa, año)
    if df_actual.empty or not meses:
        return {"rows": [], "mes": mes}
    año_ant = int(año) - 1
    df_anterior, _ = _load_and_build(empresa, str(año_ant))

    rows = []
    for _, row in df_actual.iterrows():
        cat = row["Categoria"]
        concepto = row["Concepto"] or ""
        actual_val = row.get(f"res_{mes}", 0) or 0
        anterior_val = 0
        if not df_anterior.empty:
            match = df_anterior[df_anterior["Categoria"] == cat]
            if not match.empty:
                anterior_val = match.iloc[0].get(f"res_{mes}", 0) or 0
        dif = actual_val - anterior_val
        variacion = (dif / abs(anterior_val) * 100) if anterior_val != 0 else None
        rows.append({
            "Categoria": cat,
            "Concepto": concepto,
            "is_total": row["is_total"],
            "actual": actual_val,
            "anterior": anterior_val,
            "diferencia": dif,
            "variacion": variacion,
        })
    return {"rows": rows, "mes": mes}


def build_excel_bytes(empresa: str, año: str):
    df_rows, meses = _load_and_build(empresa, año)
    if df_rows.empty:
        return None
    return _build_excel(df_rows, meses)
