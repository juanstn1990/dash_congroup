"""Balance P&G dashboard page – Comparación P y G."""
import base64
import io
import re

import dash_bootstrap_components as dbc
import duckdb
import pandas as pd
import plotly.graph_objects as go
from dash import Input, Output, State, callback, dcc, html, no_update

DB_PATH = "mazda.duckdb"

AZUL            = "#1B3A6B"
AZUL_ALT        = "#254E9A"   # header meses impares
ROJO            = "#C0392B"
VERDE           = "#0D7C66"
GRIS_TOTAL      = "#2D3748"
COL_ODD         = "#EEF3FB"   # fondo dato mes impar, fila par
COL_ODD_STRIPE  = "#E4EDF8"   # fondo dato mes impar, fila impar
COL_TOT         = "#E0EAF8"   # columna Total, fila par
COL_TOT_STRIPE  = "#D6E4F5"   # columna Total, fila impar
ROW_STRIPE      = "#F7F9FC"   # mes par, fila impar
CAT_W           = "80px"      # ancho columna Categoría (sticky)
CON_W           = "220px"     # ancho columna Concepto  (sticky)

MESES_LABELS = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


# ── Data helpers ──────────────────────────────────────────────────────────────

def _con():
    return duckdb.connect(DB_PATH, read_only=True)


def get_filter_options():
    con = _con()
    empresas = con.execute(
        'SELECT DISTINCT empresa FROM mazda.main.dim_empresa ORDER BY 1'
    ).fetchdf()["empresa"].tolist()
    años = con.execute(
        """
        SELECT DISTINCT dt.año
        FROM mazda.main.dim_tiempo dt
        WHERE EXISTS (
            SELECT 1 FROM mazda.main.balance b WHERE CAST(b."Año" AS INT) = dt.año
        )
        ORDER BY 1 DESC
        """
    ).fetchdf()["año"].tolist()
    con.close()
    return empresas, años


def _parse_denom_cats(porcentaje: str) -> list:
    """Extract denominator category codes from a Porcentaje formula string."""
    if not porcentaje or pd.isna(porcentaje):
        return []
    parts = porcentaje.split("/", 1)
    if len(parts) < 2:
        return []
    return re.findall(r"\b([A-Z][A-Z0-9]*\d+)\b", parts[1])


def _load_and_build(empresa, año):
    con = _con()

    conditions = []
    if empresa and empresa != "Todas":
        conditions.append(f'"Empresa Nombre" = \'{empresa}\'')
    if año and año != "Todos":
        conditions.append(f'"Año" = {int(año)}')
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    df_leaf = con.execute(f"""
        SELECT Clave, Mes, SUM("Saldo Mensual") AS Resultado
        FROM mazda.main.balance
        {where}
        GROUP BY Clave, Mes
    """).fetchdf()

    plantilla = con.execute(
        "SELECT * FROM mazda.main.plantilla_resultado ORDER BY N"
    ).fetchdf()
    con.close()

    if df_leaf.empty:
        return pd.DataFrame(), []

    # Build cat_values: {cat: {mes: value}}
    cat_values: dict[str, dict] = {}
    for _, row in df_leaf.iterrows():
        mes = int(row["Mes"]) if not pd.isna(row["Mes"]) else None
        if mes is None:
            continue
        cat_values.setdefault(row["Clave"], {})[mes] = row["Resultado"]

    meses = sorted({m for cv in cat_values.values() for m in cv})

    # Compute T-row totals bottom-up (plantilla is already ordered by N)
    for _, t_row in plantilla[
        plantilla["Categoria"].str.startswith("T", na=False)
    ].iterrows():
        t_cat = t_row["Categoria"]
        children = plantilla[plantilla["GrupoPadre"] == t_cat]["Categoria"].tolist()
        cat_values.setdefault(t_cat, {})
        for m in meses:
            cat_values[t_cat][m] = sum(
                cat_values.get(c, {}).get(m, 0) or 0 for c in children
            )

    # Add cross-month total
    for cat in cat_values:
        cat_values[cat]["Total"] = sum(cat_values[cat].get(m, 0) or 0 for m in meses)

    # Build display rows in plantilla order
    cols = meses + ["Total"]
    rows = []
    for _, pr in plantilla.iterrows():
        cat = pr["Categoria"]
        denom_cats = _parse_denom_cats(pr["Porcentaje"])
        vals = cat_values.get(cat, {})

        r: dict = {
            "Categoria": cat,
            "Concepto": pr["Concepto"],
            "is_total": str(cat).startswith("T"),
        }
        for col in cols:
            val = vals.get(col, 0) or 0
            denom = sum(cat_values.get(c, {}).get(col, 0) or 0 for c in denom_cats)
            r[f"res_{col}"] = val
            r[f"pct_{col}"] = (val / denom * 100) if denom != 0 else None
        rows.append(r)

    return pd.DataFrame(rows), meses


# ── Rendering helpers ─────────────────────────────────────────────────────────

def _fmt(v, pct=False) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if pct:
        return f"{v:.2f} %"
    rounded = int(round(v))
    if rounded < 0:
        return f"({abs(rounded):,})".replace(",", ".")
    return f"{rounded:,}".replace(",", ".")


def _th(text, colspan=1, rowspan=1, center=False, top="0px", bg=None, left=None):
    style = {
        "padding": "6px 10px",
        "background": bg or AZUL,
        "color": "white",
        "fontWeight": "600",
        "textAlign": "center" if center else "left",
        "verticalAlign": "middle",
        "boxShadow": "inset 0 0 0 1px #FFFFFF",
        "fontSize": "12px",
        "whiteSpace": "nowrap",
        "position": "sticky",
        "top": top,
        "zIndex": "5" if left is not None else "3",
    }
    if left is not None:
        style["left"] = left
    return html.Th(text, colSpan=colspan, rowSpan=rowspan, style=style)


def _td(text, bg="white", fg="#1C2B3A", right=False, bold=False,
        left=None, border_left=False, indent=False):
    # White solid borders for row-header (sticky) cells, grey for regular data cells
    is_row_header = left is not None
    style = {
        "padding": "5px 10px",
        "background": bg,
        "color": fg,
        "textAlign": "right" if right else "left",
        "borderTop": "1px solid #FFFFFF" if is_row_header else "1px solid #E2E8F0",
        "borderBottom": "1px solid #FFFFFF" if is_row_header else "1px solid #E2E8F0",
        "borderRight": "1px solid #FFFFFF" if is_row_header else "1px solid #E2E8F0",
        "borderLeft": "3px solid #8FA8C8" if border_left else ("1px solid #FFFFFF" if is_row_header else "1px solid #E2E8F0"),
        "fontWeight": "bold" if bold else "normal",
        "fontSize": "13px",
        "whiteSpace": "nowrap",
        "paddingLeft": "20px" if indent else "10px",
    }
    if is_row_header:
        style["position"] = "sticky"
        style["left"] = left
        style["zIndex"] = "2"
    return html.Td(text, style=style)


def _build_table(df_rows, meses):
    if df_rows is None or df_rows.empty:
        return html.Div(
            "No hay datos para los filtros seleccionados.",
            style={"padding": "40px", "textAlign": "center", "color": "#888"},
        )

    cols = meses + ["Total"]

    def _hdr_bg(i):
        if i >= len(meses):
            return "#163266"
        return AZUL_ALT if i % 2 == 1 else AZUL

    def _data_bg(col_i, row_i, is_total_row):
        if is_total_row:
            return GRIS_TOTAL
        stripe = row_i % 2 == 1
        if col_i >= len(meses):
            return COL_TOT_STRIPE if stripe else COL_TOT
        if col_i % 2 == 1:
            return COL_ODD_STRIPE if stripe else COL_ODD
        return ROW_STRIPE if stripe else "white"

    def _sticky_bg(is_total_row, row_i):
        if is_total_row:
            return GRIS_TOTAL
        return ROW_STRIPE if row_i % 2 == 1 else "white"

    # ── Header row 1 ─────────────────────────────────────────────────────────
    hdr1 = [
        _th("Categoría", top="0px", left="0px", rowspan=2),
        _th("Concepto",  top="0px", left=CAT_W, rowspan=2),
    ]
    for i, m in enumerate(meses):
        hdr1.append(_th(MESES_LABELS.get(m, str(m)), colspan=2, center=True, bg=_hdr_bg(i)))
    hdr1.append(_th("Total", colspan=2, center=True, bg=_hdr_bg(len(meses))))

    # ── Header row 2 (sin celdas para Categoría/Concepto — cubiertas por rowSpan) ──
    hdr2 = []
    for i in range(len(cols)):
        hdr2.append(_th("Resultado", center=True, top="32px", bg=_hdr_bg(i)))
        hdr2.append(_th("%",         center=True, top="32px", bg=_hdr_bg(i)))

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_rows = []
    for row_i, (_, row) in enumerate(df_rows.iterrows()):
        is_total = row["is_total"]
        fg       = "white" if is_total else "#1C2B3A"
        s_bg     = _sticky_bg(is_total, row_i)

        cells = [
            _td(row["Categoria"] or "", bg=s_bg, fg=fg, bold=True,  left="0px"),
            _td(row["Concepto"]  or "", bg=s_bg, fg=fg, bold=is_total,
                left=CAT_W, indent=not is_total),
        ]
        for col_i, col in enumerate(cols):
            bg  = _data_bg(col_i, row_i, is_total)
            val = row.get(f"res_{col}", 0) or 0
            pct = row.get(f"pct_{col}")

            # valor: rojo si negativo (formato contable), verde si positivo en no-total
            val_fg = (fg if is_total
                      else ROJO if val < 0
                      else "#1C2B3A")

            # porcentaje: verde/rojo según signo
            pct_fg = (fg if is_total
                      else VERDE if (pct is not None and not pd.isna(pct) and pct > 0)
                      else ROJO  if (pct is not None and not pd.isna(pct) and pct < 0)
                      else "#888")

            cells.append(_td(_fmt(val), bg=bg, fg=val_fg, right=True,
                             bold=is_total, border_left=(col_i < len(meses))))
            cells.append(_td(_fmt(pct, pct=True), bg=bg, fg=pct_fg, right=True))

        data_rows.append(html.Tr(cells))

    return html.Div(
        html.Table(
            [
                html.Thead([html.Tr(hdr1), html.Tr(hdr2)]),
                html.Tbody(data_rows),
            ],
            style={
                "width": "100%",
                "borderCollapse": "collapse",
                "fontSize": "13px",
                "fontFamily": "Inter, sans-serif",
            },
        ),
        style={"overflowX": "auto", "overflowY": "auto", "maxHeight": "calc(100vh - 210px)"},
    )


# ── Resultado vs Presupuesto ──────────────────────────────────────────────────

def _load_and_build_vs_ppto(empresa, año):
    df_real, meses_real = _load_and_build(empresa, año)
    df_ppto, meses_ppto = _load_and_build_ppto(empresa, año)

    if df_real.empty and df_ppto.empty:
        return pd.DataFrame(), []

    meses = sorted(set(meses_real) | set(meses_ppto))
    cols = meses + ["Total"]

    ppto_lookup = {row["Categoria"]: row for _, row in df_ppto.iterrows()} if not df_ppto.empty else {}

    rows = []
    base = df_real if not df_real.empty else df_ppto
    for _, row in base.iterrows():
        cat = row["Categoria"]
        ppto_row = ppto_lookup.get(cat, {})
        r = {
            "Categoria": cat,
            "Concepto": row["Concepto"],
            "is_total": row["is_total"],
        }
        for col in cols:
            real_val = (row.get(f"res_{col}", 0) or 0) if not df_real.empty else 0
            ppto_val = ppto_row.get(f"ppto_{col}", 0) or 0
            diff = real_val - ppto_val
            var_pct = (diff / abs(ppto_val) * 100) if ppto_val != 0 else None
            r[f"real_{col}"] = real_val
            r[f"ppto_{col}"] = ppto_val
            r[f"diff_{col}"] = diff
            r[f"var_{col}"] = var_pct
        rows.append(r)

    return pd.DataFrame(rows), meses


def _build_vs_ppto_table(df_rows, meses):
    if df_rows is None or df_rows.empty:
        return html.Div(
            "No hay datos para los filtros seleccionados.",
            style={"padding": "40px", "textAlign": "center", "color": "#888"},
        )

    cols = meses + ["Total"]

    def _hdr_bg(i):
        if i >= len(meses):
            return "#163266"
        return AZUL_ALT if i % 2 == 1 else AZUL

    def _data_bg(col_i, row_i, is_total_row):
        if is_total_row:
            return GRIS_TOTAL
        stripe = row_i % 2 == 1
        if col_i >= len(meses):
            return COL_TOT_STRIPE if stripe else COL_TOT
        if col_i % 2 == 1:
            return COL_ODD_STRIPE if stripe else COL_ODD
        return ROW_STRIPE if stripe else "white"

    def _sticky_bg(is_total_row, row_i):
        if is_total_row:
            return GRIS_TOTAL
        return ROW_STRIPE if row_i % 2 == 1 else "white"

    # ── Header row 1 ──────────────────────────────────────────────────────────
    hdr1 = [
        _th("Categoría", top="0px", left="0px", rowspan=2),
        _th("Concepto",  top="0px", left=CAT_W, rowspan=2),
    ]
    for i, m in enumerate(meses):
        hdr1.append(_th(MESES_LABELS.get(m, str(m)), colspan=4, center=True, bg=_hdr_bg(i)))
    hdr1.append(_th("Total", colspan=4, center=True, bg=_hdr_bg(len(meses))))

    # ── Header row 2 ──────────────────────────────────────────────────────────
    hdr2 = []
    for i in range(len(cols)):
        bg = _hdr_bg(i)
        hdr2.append(_th("Real",            center=True, top="32px", bg=bg))
        hdr2.append(_th("Presupuesto",     center=True, top="32px", bg=bg))
        hdr2.append(_th("Real - Ppto",     center=True, top="32px", bg=bg))
        hdr2.append(_th("% Variación",     center=True, top="32px", bg=bg))

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_rows = []
    for row_i, (_, row) in enumerate(df_rows.iterrows()):
        is_total = row["is_total"]
        fg       = "white" if is_total else "#1C2B3A"
        s_bg     = _sticky_bg(is_total, row_i)

        cells = [
            _td(row["Categoria"] or "", bg=s_bg, fg=fg, bold=True, left="0px"),
            _td(row["Concepto"]  or "", bg=s_bg, fg=fg, bold=is_total,
                left=CAT_W, indent=not is_total),
        ]
        for col_i, col in enumerate(cols):
            bg = _data_bg(col_i, row_i, is_total)

            real_val = row.get(f"real_{col}", 0) or 0
            ppto_val = row.get(f"ppto_{col}", 0) or 0
            diff_val = row.get(f"diff_{col}", 0) or 0
            var_pct  = row.get(f"var_{col}")

            real_fg = (fg if is_total else ROJO if real_val < 0 else "#1C2B3A")
            ppto_fg = (fg if is_total else ROJO if ppto_val < 0 else "#1C2B3A")
            diff_fg = (fg if is_total
                       else VERDE if diff_val > 0
                       else ROJO  if diff_val < 0
                       else "#1C2B3A")
            var_fg  = (fg if is_total
                       else VERDE if (var_pct is not None and not pd.isna(var_pct) and var_pct > 0)
                       else ROJO  if (var_pct is not None and not pd.isna(var_pct) and var_pct < 0)
                       else "#888")

            cells.append(_td(_fmt(real_val), bg=bg, fg=real_fg, right=True,
                              bold=is_total, border_left=(col_i < len(meses))))
            cells.append(_td(_fmt(ppto_val), bg=bg, fg=ppto_fg, right=True, bold=is_total))
            cells.append(_td(_fmt(diff_val), bg=bg, fg=diff_fg, right=True, bold=is_total))
            cells.append(_td(_fmt(var_pct, pct=True), bg=bg, fg=var_fg, right=True))

        data_rows.append(html.Tr(cells))

    return html.Div(
        html.Table(
            [
                html.Thead([html.Tr(hdr1), html.Tr(hdr2)]),
                html.Tbody(data_rows),
            ],
            style={
                "width": "100%",
                "borderCollapse": "collapse",
                "fontSize": "13px",
                "fontFamily": "Inter, sans-serif",
            },
        ),
        style={"overflowX": "auto", "overflowY": "auto", "maxHeight": "calc(100vh - 210px)"},
    )


# ── Presupuesto ───────────────────────────────────────────────────────────────

def _load_and_build_ppto(empresa, año):
    con = _con()

    empresa_code = None
    if empresa and empresa != "Todas":
        row = con.execute(
            "SELECT code FROM mazda.main.dim_empresa WHERE empresa = ?", [empresa]
        ).fetchone()
        if row:
            empresa_code = int(row[0])

    conditions = []
    if empresa_code is not None:
        conditions.append(f"Empresa = {empresa_code}")
    if año and año != "Todos":
        conditions.append(f"Año = '{año}'")
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    df_leaf = con.execute(f"""
        SELECT Categoria, CAST(Mes AS INT) AS Mes, SUM(TotalBase) AS TotalBase
        FROM mazda.main.ppto_mensual
        {where}
        GROUP BY Categoria, Mes
    """).fetchdf()

    plantilla = con.execute(
        "SELECT * FROM mazda.main.plantilla_resultado ORDER BY N"
    ).fetchdf()
    con.close()

    if df_leaf.empty:
        return pd.DataFrame(), []

    cat_values: dict[str, dict] = {}
    for _, row in df_leaf.iterrows():
        mes = int(row["Mes"])
        cat_values.setdefault(row["Categoria"], {})[mes] = row["TotalBase"]

    meses = sorted({m for cv in cat_values.values() for m in cv})

    for _, t_row in plantilla[
        plantilla["Categoria"].str.startswith("T", na=False)
    ].iterrows():
        t_cat = t_row["Categoria"]
        children = plantilla[plantilla["GrupoPadre"] == t_cat]["Categoria"].tolist()
        cat_values.setdefault(t_cat, {})
        for m in meses:
            cat_values[t_cat][m] = sum(
                cat_values.get(c, {}).get(m, 0) or 0 for c in children
            )

    for cat in cat_values:
        cat_values[cat]["Total"] = sum(cat_values[cat].get(m, 0) or 0 for m in meses)

    cols = meses + ["Total"]
    rows = []
    for _, pr in plantilla.iterrows():
        cat = pr["Categoria"]
        denom_cats = _parse_denom_cats(pr["Porcentaje"])
        vals = cat_values.get(cat, {})

        r: dict = {
            "Categoria": cat,
            "Concepto": pr["Concepto"],
            "is_total": str(cat).startswith("T"),
        }
        for col in cols:
            val = vals.get(col, 0) or 0
            denom = sum(cat_values.get(c, {}).get(col, 0) or 0 for c in denom_cats)
            r[f"ppto_{col}"] = val
            r[f"pct_{col}"] = (val / denom * 100) if denom != 0 else None
        rows.append(r)

    return pd.DataFrame(rows), meses


def _build_ppto_table(df_rows, meses):
    if df_rows is None or df_rows.empty:
        return html.Div(
            "No hay datos para los filtros seleccionados.",
            style={"padding": "40px", "textAlign": "center", "color": "#888"},
        )

    cols = meses + ["Total"]

    def _hdr_bg(i):
        if i >= len(meses):
            return "#163266"
        return AZUL_ALT if i % 2 == 1 else AZUL

    def _data_bg(col_i, row_i, is_total_row):
        if is_total_row:
            return GRIS_TOTAL
        stripe = row_i % 2 == 1
        if col_i >= len(meses):
            return COL_TOT_STRIPE if stripe else COL_TOT
        if col_i % 2 == 1:
            return COL_ODD_STRIPE if stripe else COL_ODD
        return ROW_STRIPE if stripe else "white"

    def _sticky_bg(is_total_row, row_i):
        if is_total_row:
            return GRIS_TOTAL
        return ROW_STRIPE if row_i % 2 == 1 else "white"

    hdr1 = [
        _th("Categoría", top="0px", left="0px", rowspan=2),
        _th("Concepto",  top="0px", left=CAT_W, rowspan=2),
    ]
    for i, m in enumerate(meses):
        hdr1.append(_th(MESES_LABELS.get(m, str(m)), colspan=2, center=True, bg=_hdr_bg(i)))
    hdr1.append(_th("Total", colspan=2, center=True, bg=_hdr_bg(len(meses))))

    hdr2 = []
    for i in range(len(cols)):
        hdr2.append(_th("Presupuesto", center=True, top="32px", bg=_hdr_bg(i)))
        hdr2.append(_th("%",           center=True, top="32px", bg=_hdr_bg(i)))

    data_rows = []
    for row_i, (_, row) in enumerate(df_rows.iterrows()):
        is_total = row["is_total"]
        fg       = "white" if is_total else "#1C2B3A"
        s_bg     = _sticky_bg(is_total, row_i)

        cells = [
            _td(row["Categoria"] or "", bg=s_bg, fg=fg, bold=True,  left="0px"),
            _td(row["Concepto"]  or "", bg=s_bg, fg=fg, bold=is_total,
                left=CAT_W, indent=not is_total),
        ]
        for col_i, col in enumerate(cols):
            bg  = _data_bg(col_i, row_i, is_total)
            val = row.get(f"ppto_{col}", 0) or 0
            pct = row.get(f"pct_{col}")

            val_fg = (fg if is_total
                      else ROJO if val < 0
                      else "#1C2B3A")

            pct_fg = (fg if is_total
                      else VERDE if (pct is not None and not pd.isna(pct) and pct > 0)
                      else ROJO  if (pct is not None and not pd.isna(pct) and pct < 0)
                      else "#888")

            cells.append(_td(_fmt(val), bg=bg, fg=val_fg, right=True,
                             bold=is_total, border_left=(col_i < len(meses))))
            cells.append(_td(_fmt(pct, pct=True), bg=bg, fg=pct_fg, right=True))

        data_rows.append(html.Tr(cells))

    return html.Div(
        html.Table(
            [
                html.Thead([html.Tr(hdr1), html.Tr(hdr2)]),
                html.Tbody(data_rows),
            ],
            style={
                "width": "100%",
                "borderCollapse": "collapse",
                "fontSize": "13px",
                "fontFamily": "Inter, sans-serif",
            },
        ),
        style={"overflowX": "auto", "overflowY": "auto", "maxHeight": "calc(100vh - 210px)"},
    )


# ── Charts ───────────────────────────────────────────────────────────────────

def _kpi_card(label, value, color, subtitle=None):
    fmt_val = _fmt(value)
    sign_color = VERDE if value > 0 else ROJO if value < 0 else "#888"
    sign = "▲" if value > 0 else "▼" if value < 0 else "─"
    return html.Div([
        html.P(label, style={"margin": "0 0 4px 0", "fontSize": "11px",
                              "color": "#64748B", "fontWeight": "600", "textTransform": "uppercase"}),
        html.H3(fmt_val, style={"margin": "0 0 2px 0", "fontSize": "22px",
                                 "color": color, "fontWeight": "700"}),
        html.Span(f"{sign} {subtitle}" if subtitle else sign,
                  style={"fontSize": "11px", "color": sign_color, "fontWeight": "600"}),
    ], style={
        "background": "white", "borderRadius": "10px",
        "boxShadow": "0 2px 8px rgba(0,0,0,0.08)",
        "padding": "14px 18px", "borderLeft": f"4px solid {color}",
        "flex": "1",
    })


def _chart_card(fig):
    return html.Div(
        dcc.Graph(figure=fig, config={"displayModeBar": False}),
        style={"background": "white", "borderRadius": "10px",
               "boxShadow": "0 2px 8px rgba(0,0,0,0.08)", "padding": "8px"},
    )


_PLOT_BASE = dict(
    plot_bgcolor="white", paper_bgcolor="white",
    font=dict(family="Inter, sans-serif", size=12),
    margin=dict(l=60, r=30, t=50, b=40),
    yaxis=dict(gridcolor="#E8ECF2", zeroline=True, zerolinecolor="#C0CCDC"),
    xaxis=dict(gridcolor="#E8ECF2"),
    legend=dict(orientation="h", y=1.12),
)


def _build_kpis(df_rows, meses):
    if df_rows is None or df_rows.empty or not meses:
        return html.Div()

    t = {row["Categoria"]: row for _, row in df_rows[df_rows["is_total"]].iterrows()}

    def tv(cat, col="Total"):
        return t.get(cat, {}).get(f"res_{col}", 0) or 0

    pyg_total = tv("T18")
    bn_total  = tv("T13")
    gastos_g  = tv("T14")
    gastos_p  = tv("T15")

    pyg_by_mes = {m: tv("T18", m) for m in meses}
    mejor_m    = max(pyg_by_mes, key=pyg_by_mes.get)
    peor_m     = min(pyg_by_mes, key=pyg_by_mes.get)

    return html.Div([
        _kpi_card("PyG Total",          pyg_total, VERDE if pyg_total >= 0 else ROJO),
        _kpi_card("BN Total",           bn_total,  AZUL),
        _kpi_card("Gastos Generales",   abs(gastos_g), ROJO),
        _kpi_card("Gastos Personal",    abs(gastos_p), ROJO),
        _kpi_card("Mejor Mes",          pyg_by_mes[mejor_m], VERDE,
                  MESES_LABELS[mejor_m].capitalize()),
        _kpi_card("Peor Mes",           pyg_by_mes[peor_m],  ROJO,
                  MESES_LABELS[peor_m].capitalize()),
    ], style={"display": "flex", "gap": "12px", "marginBottom": "16px", "flexWrap": "wrap"})


def _build_charts(df_rows, meses):
    if df_rows is None or df_rows.empty or not meses:
        return html.Div()

    mes_labels = [MESES_LABELS.get(m, str(m)).capitalize() for m in meses]

    # Quick lookups: T-rows indexed by Categoria
    t = {row["Categoria"]: row for _, row in df_rows[df_rows["is_total"]].iterrows()}

    def tv(cat, col="Total"):
        return t.get(cat, {}).get(f"res_{col}", 0) or 0

    def tm(cat):
        return [tv(cat, m) for m in meses]

    def pm(cat):
        return [t.get(cat, {}).get(f"pct_{m}") for m in meses]

    # ── Chart 1: Evolución mensual PyG vs BN Total ─────────────────────────
    pyg_mes = tm("T18")
    bn_mes  = tm("T13")

    fig1 = go.Figure()
    fig1.add_trace(go.Bar(
        x=mes_labels, y=pyg_mes, name="PyG",
        marker_color=[VERDE if v >= 0 else ROJO for v in pyg_mes],
        text=[_fmt(v) for v in pyg_mes],
        textposition="outside", textfont=dict(size=10),
    ))
    fig1.add_trace(go.Scatter(
        x=mes_labels, y=bn_mes, name="BN Total",
        mode="lines+markers",
        line=dict(color=AZUL, width=2, dash="dot"),
        marker=dict(size=7),
    ))
    fig1.update_layout(title="Evolución mensual: PyG vs BN Total", height=320, **_PLOT_BASE)

    # ── Chart 2: BN por área de negocio ───────────────────────────────────
    fig2 = go.Figure()
    for label, cat, color in [
        ("BN VN",      "T2",  "#1B3A6B"),
        ("BN VO",      "T4",  "#2E6BE6"),
        ("BN Posventa","T12", "#0D7C66"),
    ]:
        vals = tm(cat)
        fig2.add_trace(go.Bar(
            x=mes_labels, y=vals, name=label, marker_color=color,
            text=[_fmt(v) for v in vals],
            textposition="inside", textfont=dict(size=10, color="white"),
        ))
    fig2.update_layout(title="BN por Área de Negocio", barmode="group",
                       height=320, **_PLOT_BASE)

    # ── Chart 3: Cascada P&G (waterfall) ──────────────────────────────────
    t13, t14, t15 = tv("T13"), tv("T14"), tv("T15")
    t16, t17, t18 = tv("T16"), tv("T17"), tv("T18")

    wf_x = ["BN Total", "G.Generales", "G.Personal",
             "Res. I+A", "I+Amort.", "Res. Op.", "Otros", "PyG"]
    wf_y = [t13, t14, t15, t16, t17 - t16, t17, t18 - t17, t18]
    wf_m = ["absolute", "relative", "relative",
             "total", "relative", "total", "relative", "total"]

    fig3 = go.Figure(go.Waterfall(
        orientation="v", measure=wf_m, x=wf_x, y=wf_y,
        connector=dict(line=dict(color="#B0BEC5", width=1)),
        increasing=dict(marker=dict(color=VERDE)),
        decreasing=dict(marker=dict(color=ROJO)),
        totals=dict(marker=dict(color=AZUL)),
        text=[_fmt(v) for v in wf_y],
        textposition="outside", textfont=dict(size=10),
    ))
    fig3.update_layout(title="Cascada del Resultado (PyG)", height=360,
                       **_PLOT_BASE)

    # ── Chart 4: Evolución de márgenes % ──────────────────────────────────
    fig4 = go.Figure()
    for label, cat, color in [
        ("BB VN %",    "T1",  "#1B3A6B"),
        ("BN VN %",    "T2",  "#2E6BE6"),
        ("BN Total %", "T13", "#0D7C66"),
        ("PyG %",      "T18", "#E07B39"),
    ]:
        vals = pm(cat)
        if any(v is not None for v in vals):
            fig4.add_trace(go.Scatter(
                x=mes_labels, y=vals, name=label,
                mode="lines+markers",
                line=dict(color=color, width=2),
                marker=dict(size=7),
                hovertemplate=f"<b>{label}</b><br>%{{x}}: %{{y:.2f}} %<extra></extra>",
            ))
    fig4.add_hline(y=0, line_dash="dot", line_color=ROJO, line_width=1)
    fig4.update_layout(
        title="Evolución de Márgenes (%)",
        height=320,
        yaxis=dict(gridcolor="#E8ECF2", zeroline=True, zerolinecolor="#C0CCDC", ticksuffix=" %"),
        **{k: v for k, v in _PLOT_BASE.items() if k != "yaxis"},
    )

    # ── Layout final ──────────────────────────────────────────────────────
    return html.Div([
        html.H4("Análisis Contable", style={
            "color": AZUL, "fontWeight": "700",
            "marginBottom": "16px", "marginTop": "28px",
            "borderBottom": f"2px solid {AZUL}", "paddingBottom": "8px",
        }),
        html.Div([
            html.Div(_chart_card(fig1), style={"flex": "1", "minWidth": 0}),
            html.Div(_chart_card(fig4), style={"flex": "1", "minWidth": 0}),
        ], style={"display": "flex", "gap": "16px", "marginBottom": "16px"}),
        html.Div([
            html.Div(_chart_card(fig2), style={"flex": "2", "minWidth": 0}),
            html.Div(_chart_card(fig3), style={"flex": "3", "minWidth": 0}),
        ], style={"display": "flex", "gap": "16px"}),
    ])


# ── Resultado vs Ejercicio Anterior ───────────────────────────────────────────

def _build_vs_table(df_actual, df_anterior, mes):
    if df_actual is None or df_actual.empty:
        return html.Div(
            "No hay datos para comparar.",
            style={"padding": "40px", "textAlign": "center", "color": "#888"},
        )

    hdr = html.Tr([
        _th("Concepto", center=False),
        _th("Resultado Actual", center=True),
        _th("Resultado Anterior", center=True),
        _th("Diferencia", center=True),
        _th("Variación %", center=True),
    ])

    data_rows = []
    for row_i, (_, row) in enumerate(df_actual.iterrows()):
        is_total = row["is_total"]
        concepto = row["Concepto"] or ""
        actual_val = row.get(f"res_{mes}", 0) or 0

        anterior_val = 0
        if not df_anterior.empty:
            match = df_anterior[df_anterior["Categoria"] == row["Categoria"]]
            if not match.empty:
                anterior_val = match.iloc[0].get(f"res_{mes}", 0) or 0

        dif = actual_val - anterior_val
        variacion = (dif / abs(anterior_val) * 100) if anterior_val != 0 else None

        fg = "white" if is_total else "#1C2B3A"
        bg = GRIS_TOTAL if is_total else ("#F7F9FC" if row_i % 2 == 1 else "white")

        actual_fg = fg if is_total else (ROJO if actual_val < 0 else "#1C2B3A")
        anterior_fg = fg if is_total else (ROJO if anterior_val < 0 else "#1C2B3A")
        dif_fg = fg if is_total else (ROJO if dif < 0 else VERDE if dif > 0 else "#1C2B3A")
        var_fg = fg if is_total else (
            ROJO if (variacion is not None and variacion < 0)
            else VERDE if (variacion is not None and variacion > 0)
            else "#888"
        )

        cells = [
            _td(concepto, bg=bg, fg=fg, bold=is_total, indent=not is_total),
            _td(_fmt(actual_val), bg=bg, fg=actual_fg, right=True, bold=is_total),
            _td(_fmt(anterior_val), bg=bg, fg=anterior_fg, right=True, bold=is_total),
            _td(_fmt(dif), bg=bg, fg=dif_fg, right=True, bold=is_total),
            _td(_fmt(variacion, pct=True), bg=bg, fg=var_fg, right=True, bold=is_total),
        ]
        data_rows.append(html.Tr(cells))

    return html.Div(
        html.Table(
            [html.Thead(hdr), html.Tbody(data_rows)],
            style={
                "width": "100%",
                "borderCollapse": "collapse",
                "fontSize": "13px",
                "fontFamily": "Inter, sans-serif",
            },
        ),
        style={"overflowX": "auto", "overflowY": "auto", "maxHeight": "calc(100vh - 210px)"},
    )


# ── Excel export ──────────────────────────────────────────────────────────────

def _build_excel(df_rows, meses):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    cols = meses + ["Total"]

    export_cols = ["Categoría", "Concepto"]
    for col in cols:
        label = MESES_LABELS.get(col, str(col)).capitalize() if isinstance(col, int) else col
        export_cols.append(f"{label} Resultado")
        export_cols.append(f"{label} %")

    data = []
    for _, row in df_rows.iterrows():
        r = [row["Categoria"], row["Concepto"]]
        for col in cols:
            r.append(row.get(f"res_{col}", 0) or 0)
            r.append(row.get(f"pct_{col}"))
        data.append(r)

    df_export = pd.DataFrame(data, columns=export_cols)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Balance P&G")
        ws = writer.sheets["Balance P&G"]

        # Paleta de estilos
        header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        total_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        total_font = Font(color="FFFFFF", bold=True, size=11)
        odd_fill = PatternFill(start_color="EEF3FB", end_color="EEF3FB", fill_type="solid")
        odd_stripe_fill = PatternFill(start_color="E4EDF8", end_color="E4EDF8", fill_type="solid")
        tot_col_fill = PatternFill(start_color="E0EAF8", end_color="E0EAF8", fill_type="solid")
        tot_col_stripe_fill = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")
        stripe_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
        red_font = Font(color="C0392B")
        green_font = Font(color="0D7C66")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Anchos
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 38
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col_letter not in ("A", "B"):
                ws.column_dimensions[col_letter].width = 18

        # Filas de datos
        for row_idx, (_, row) in enumerate(df_rows.iterrows()):
            excel_row = row_idx + 2
            is_total = row["is_total"]
            stripe = row_idx % 2 == 1

            for col_idx, cell in enumerate(ws[excel_row], start=1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="right" if col_idx > 2 else "left", vertical="center"
                )

                if is_total:
                    cell.fill = total_fill
                    cell.font = total_font
                    continue

                # Fondo según columna / stripe
                if col_idx > 2:
                    if col_idx > len(export_cols) - 2:   # columnas Total
                        cell.fill = tot_col_stripe_fill if stripe else tot_col_fill
                    else:
                        mes_col_idx = (col_idx - 3) // 2
                        if mes_col_idx % 2 == 1:
                            cell.fill = odd_stripe_fill if stripe else odd_fill
                        else:
                            cell.fill = stripe_fill if stripe else PatternFill(fill_type=None)
                else:
                    cell.fill = stripe_fill if stripe else PatternFill(fill_type=None)

                # Color de fuente numérica
                if col_idx > 2:
                    val = cell.value
                    if val is not None and isinstance(val, (int, float)) and not pd.isna(val):
                        if col_idx % 2 == 0:   # columna %
                            cell.font = green_font if val > 0 else red_font if val < 0 else Font()
                        else:                  # columna Resultado
                            cell.font = red_font if val < 0 else Font()

            # Formatos numéricos
            for col_idx in range(3, len(export_cols) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                if col_idx % 2 == 0:
                    cell.number_format = '0.00"%"'
                else:
                    cell.number_format = '#,##0;(#,##0)'

        ws.freeze_panes = "C2"

    buffer.seek(0)
    return buffer.getvalue()


# ── Public layout ─────────────────────────────────────────────────────────────

def balance_content():
    """Returns the content div (without sidebar). Call once at startup."""
    empresas, años = get_filter_options()
    año_default = max(años) if años else 2026
    empresa_default = "CANTON MOTOR" if "CANTON MOTOR" in empresas else "Todas"

    return html.Div(
        [
            dcc.Store(id="bal-data-store", data=None),
            dcc.Download(id="bal-download"),
            # Filter bar
            html.Div(
                [
                    html.Div(
                        [
                            html.Label(
                                "Empresa",
                                style={
                                    "fontWeight": "600",
                                    "color": AZUL,
                                    "fontSize": "13px",
                                    "marginBottom": "4px",
                                    "display": "block",
                                },
                            ),
                            dcc.Dropdown(
                                id="bal-empresa",
                                options=[{"label": "Todas", "value": "Todas"}]
                                + [{"label": e, "value": e} for e in empresas],
                                value=empresa_default,
                                clearable=False,
                                style={"fontSize": "13px", "minWidth": "260px"},
                            ),
                        ],
                        style={"marginRight": "20px"},
                    ),
                    html.Div(
                        [
                            html.Label(
                                "Año",
                                style={
                                    "fontWeight": "600",
                                    "color": AZUL,
                                    "fontSize": "13px",
                                    "marginBottom": "4px",
                                    "display": "block",
                                },
                            ),
                            dcc.Dropdown(
                                id="bal-año",
                                options=[{"label": "Todos", "value": "Todos"}]
                                + [{"label": str(a), "value": a} for a in años],
                                value=año_default,
                                clearable=False,
                                style={"fontSize": "13px", "minWidth": "120px"},
                            ),
                        ]
                    ),
                ],
                style={
                    "display": "flex",
                    "alignItems": "flex-end",
                    "background": "white",
                    "padding": "16px 20px",
                    "borderRadius": "8px",
                    "boxShadow": "0 2px 4px rgba(0,0,0,0.08)",
                    "marginBottom": "16px",
                    "borderLeft": f"4px solid {AZUL}",
                },
            ),
            dbc.Tabs(
                [
                    dbc.Tab(
                        label="Resultado",
                        tab_id="tab-resultado",
                        children=html.Div(
                            [
                                html.Div(id="bal-kpis", style={"marginBottom": "16px"}),
                                html.Div(
                                    [
                                        html.Div(
                                            [
                                                html.H4(
                                                    "Balance P&G",
                                                    style={
                                                        "margin": "0",
                                                        "color": AZUL,
                                                        "fontWeight": "700",
                                                        "fontSize": "16px",
                                                    },
                                                ),
                                                html.Button(
                                                    "⬇ Descargar Excel",
                                                    id="bal-download-btn",
                                                    n_clicks=0,
                                                    style={
                                                        "background": AZUL,
                                                        "color": "white",
                                                        "border": "none",
                                                        "borderRadius": "6px",
                                                        "padding": "8px 16px",
                                                        "fontSize": "13px",
                                                        "fontWeight": "600",
                                                        "cursor": "pointer",
                                                        "display": "flex",
                                                        "alignItems": "center",
                                                        "gap": "6px",
                                                    },
                                                ),
                                            ],
                                            style={
                                                "display": "flex",
                                                "justifyContent": "space-between",
                                                "alignItems": "center",
                                                "padding": "14px 20px",
                                                "borderBottom": "1px solid #E2E8F0",
                                            },
                                        ),
                                        html.Div(id="bal-table"),
                                    ],
                                    style={
                                        "background": "white",
                                        "borderRadius": "8px",
                                        "boxShadow": "0 2px 4px rgba(0,0,0,0.08)",
                                        "overflow": "hidden",
                                    },
                                ),
                                html.Div(id="bal-charts"),
                            ],
                            style={"paddingTop": "16px"},
                        ),
                    ),
                    dbc.Tab(
                        label="Resultado vs Presupuesto",
                        tab_id="tab-vs-ppto",
                        children=html.Div(
                            [
                                html.Div(
                                    [
                                        html.H4(
                                            "Resultado vs Presupuesto",
                                            style={
                                                "margin": "0",
                                                "color": AZUL,
                                                "fontWeight": "700",
                                                "fontSize": "16px",
                                            },
                                        ),
                                    ],
                                    style={
                                        "display": "flex",
                                        "justifyContent": "space-between",
                                        "alignItems": "center",
                                        "padding": "14px 20px",
                                        "borderBottom": "1px solid #E2E8F0",
                                    },
                                ),
                                html.Div(
                                    id="vs-ppto-table",
                                    style={
                                        "background": "white",
                                        "borderRadius": "8px",
                                        "boxShadow": "0 2px 4px rgba(0,0,0,0.08)",
                                        "overflow": "hidden",
                                    },
                                ),
                            ],
                            style={"paddingTop": "16px"},
                        ),
                    ),
                    dbc.Tab(
                        label="Presupuesto",
                        tab_id="tab-presupuesto",
                        children=html.Div(
                            [
                                html.Div(
                                    [
                                        html.H4(
                                            "Presupuesto",
                                            style={
                                                "margin": "0",
                                                "color": AZUL,
                                                "fontWeight": "700",
                                                "fontSize": "16px",
                                            },
                                        ),
                                    ],
                                    style={
                                        "display": "flex",
                                        "justifyContent": "space-between",
                                        "alignItems": "center",
                                        "padding": "14px 20px",
                                        "borderBottom": "1px solid #E2E8F0",
                                    },
                                ),
                                html.Div(
                                    id="ppto-table",
                                    style={
                                        "background": "white",
                                        "borderRadius": "8px",
                                        "boxShadow": "0 2px 4px rgba(0,0,0,0.08)",
                                        "overflow": "hidden",
                                    },
                                ),
                            ],
                            style={"paddingTop": "16px"},
                        ),
                    ),
                    dbc.Tab(
                        label="Resultado vs Ejercicio Anterior",
                        tab_id="tab-vs",
                        children=html.Div(
                            [
                                html.Div(
                                    [
                                        html.Label(
                                            "Mes a comparar",
                                            style={
                                                "fontWeight": "600",
                                                "color": AZUL,
                                                "fontSize": "13px",
                                                "marginBottom": "4px",
                                                "display": "block",
                                            },
                                        ),
                                        dcc.Dropdown(
                                            id="bal-vs-mes",
                                            clearable=False,
                                            style={"fontSize": "13px", "minWidth": "200px"},
                                        ),
                                    ],
                                    style={"marginBottom": "16px", "maxWidth": "260px"},
                                ),
                                html.Div(
                                    id="bal-vs-table",
                                    style={
                                        "background": "white",
                                        "borderRadius": "8px",
                                        "boxShadow": "0 2px 4px rgba(0,0,0,0.08)",
                                        "overflow": "hidden",
                                    },
                                ),
                            ],
                            style={"paddingTop": "16px"},
                        ),
                    ),
                ],
                id="bal-tabs",
                active_tab="tab-resultado",
            ),
        ]
    )


def register_callbacks():
    @callback(
        [
            Output("bal-table", "children"),
            Output("bal-charts", "children"),
            Output("bal-kpis", "children"),
            Output("bal-data-store", "data"),
        ],
        [Input("bal-empresa", "value"), Input("bal-año", "value")],
    )
    def _update(empresa, año):
        df_rows, meses = _load_and_build(empresa, año)
        store_data = None
        if not df_rows.empty:
            store_data = {
                "records": df_rows.to_dict("records"),
                "meses": meses,
            }
        return (
            _build_table(df_rows, meses),
            _build_charts(df_rows, meses),
            _build_kpis(df_rows, meses),
            store_data,
        )

    @callback(
        Output("bal-download", "data"),
        Input("bal-download-btn", "n_clicks"),
        State("bal-data-store", "data"),
        prevent_initial_call=True,
    )
    def _download(n_clicks, store_data):
        if not n_clicks or not store_data:
            return no_update
        df_rows = pd.DataFrame(store_data["records"])
        meses = store_data["meses"]
        xlsx = _build_excel(df_rows, meses)
        return dcc.send_bytes(xlsx, filename="balance_pyg.xlsx")

    @callback(
        [
            Output("bal-vs-table", "children"),
            Output("bal-vs-mes", "options"),
            Output("bal-vs-mes", "value"),
        ],
        [
            Input("bal-empresa", "value"),
            Input("bal-año", "value"),
            Input("bal-vs-mes", "value"),
        ],
    )
    def _update_vs(empresa, año, mes_seleccionado):
        if not año or año == "Todos":
            return (
                html.Div(
                    "Seleccione un año específico para comparar.",
                    style={"padding": "40px", "textAlign": "center", "color": "#888"},
                ),
                [],
                None,
            )

        df_actual, meses = _load_and_build(empresa, año)
        if df_actual.empty or not meses:
            return (
                html.Div(
                    "No hay datos para los filtros seleccionados.",
                    style={"padding": "40px", "textAlign": "center", "color": "#888"},
                ),
                [],
                None,
            )

        options = [
            {"label": MESES_LABELS.get(m, str(m)).capitalize(), "value": m}
            for m in meses
        ]

        if mes_seleccionado is None or mes_seleccionado not in meses:
            mes_seleccionado = meses[-1]

        año_ant = int(año) - 1
        df_anterior, _ = _load_and_build(empresa, año_ant)

        return (
            _build_vs_table(df_actual, df_anterior, mes_seleccionado),
            options,
            mes_seleccionado,
        )

    @callback(
        Output("ppto-table", "children"),
        [Input("bal-empresa", "value"), Input("bal-año", "value")],
    )
    def _update_ppto(empresa, año):
        df_rows, meses = _load_and_build_ppto(empresa, año)
        return _build_ppto_table(df_rows, meses)

    @callback(
        Output("vs-ppto-table", "children"),
        [Input("bal-empresa", "value"), Input("bal-año", "value")],
    )
    def _update_vs_ppto(empresa, año):
        df_rows, meses = _load_and_build_vs_ppto(empresa, año)
        return _build_vs_ppto_table(df_rows, meses)
